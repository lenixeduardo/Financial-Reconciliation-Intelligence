from __future__ import annotations

import csv
import io
from typing import Any

from openpyxl import load_workbook


def _normalize(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def parse_table(filename: str, data: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        text = data.decode("utf-8-sig")
        sample = text[:4096]
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        rows = [{str(k): _normalize(v) for k, v in row.items()} for row in reader]
        return list(reader.fieldnames or []), rows

    if lower.endswith((".xlsx", ".xlsm")):
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        values = ws.iter_rows(values_only=True)
        try:
            header = next(values)
        except StopIteration:
            return [], []
        columns = [str(v).strip() if v is not None else f"column_{i+1}" for i, v in enumerate(header)]
        rows = []
        for row in values:
            if not any(v is not None for v in row):
                continue
            rows.append({columns[i]: _normalize(row[i] if i < len(row) else None) for i in range(len(columns))})
        return columns, rows

    raise ValueError("Formato não suportado. Use CSV ou XLSX.")


def suggest_mapping(columns: list[str]) -> dict[str, str | None]:
    aliases = {
        "amount": ["valor", "amount", "vl", "total", "montante", "credito", "debito"],
        "date": ["data", "date", "dt", "lancamento", "competencia"],
        "description": ["descricao", "descrição", "description", "historico", "histórico", "cliente", "nome"],
        "document": ["documento", "document", "id", "referencia", "referência", "nf", "nota"],
    }
    lowered = {c: c.lower().strip().replace("_", " ") for c in columns}
    result: dict[str, str | None] = {k: None for k in aliases}
    for target, words in aliases.items():
        for column, normalized in lowered.items():
            if any(word in normalized for word in words):
                result[target] = column
                break
    return result
